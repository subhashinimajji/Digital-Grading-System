
from flask import Flask, render_template, request, redirect, url_for, session
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
import os
import pytesseract
import cv2
import string
from nltk.util import ngrams
from nltk.translate.bleu_score import sentence_bleu
from nltk.translate.bleu_score import SmoothingFunction
from nltk.stem import PorterStemmer
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
from sklearn.feature_extraction.text import TfidfVectorizer
import nltk
import PyPDF2
from docx import Document
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import openpyxl
import torch
import numpy as np
import spacy
import re
import torch
from transformers import AutoTokenizer, AutoModelForSequenceClassification
from nltk.translate.bleu_score import sentence_bleu, SmoothingFunction
from nltk.tokenize import word_tokenize
from sklearn.metrics import mean_squared_error
from flask import Flask

app = Flask(__name__)
app = Flask(__name__, static_url_path='/static')

app.secret_key = 'e87d32a30b00294f52a3f94f43c95674'  
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'
app.config['STATIC_FOLDER'] = 'static'
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "admin_login"

UPLOAD_FOLDER = os.path.abspath('uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Sample admin user (you can replace this with a database)
class Admin(UserMixin):
    def __init__(self, id):
        self.id = id

# Sample admin user (you can replace this with a database)
admins = {'admin': {'password': 'password'}}
def calculate_jaccard_similarity(text1, text2, n=1):
    words1 = text1.split()
    words2 = text2.split()

    ngrams1 = set(ngrams(words1, n))
    ngrams2 = set(ngrams(words2, n))

    intersection = ngrams1.intersection(ngrams2)
    union = ngrams1.union(ngrams2)

    jaccard_similarity = len(intersection) / len(union)
    return jaccard_similarity
def evaluate_answer_with_keywords(answer_text, keywords):
    # Calculate the number of overlapping keywords
    answer_words = answer_text.split()
    overlapping_keywords = [word for word in answer_words if word in keywords]
    
    # Calculate the score based on the number of overlapping keywords
    score = (len(overlapping_keywords) / len(keywords)) * 10
    return score
def extract_text_from_pdf(pdf_path):
    text = ""
    with open(pdf_path, 'rb') as pdf_file:
        pdf_reader = PyPDF2.PdfReader(pdf_file)
        for page_num in range(len(pdf_reader.pages)):
            page = pdf_reader.pages[page_num]
            text += page.extract_text()
    return text

def extract_text_from_docx(docx_path):
    doc = Document(docx_path)
    text = ""
    for paragraph in doc.paragraphs:
        text += paragraph.text + '\n'
    return text

def extract_text_from_txt(txt_path):
    with open(txt_path, 'r') as txt_file:
        text = txt_file.read()
    return text
def extract_text_from_excel(excel_path):
    text = ""
    workbook = openpyxl.load_workbook(excel_path)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    text += str(cell.value) + ' '
            text += '\n'
    return text
def extract_text_from_file(file_path):
    if file_path.lower().endswith('.pdf'):
        return extract_text_from_pdf(file_path)
    elif file_path.lower().endswith('.docx'):
        return extract_text_from_docx(file_path)
    elif file_path.lower().endswith('.txt'):
        return extract_text_from_txt(file_path)
    elif file_path.lower().endswith('.xlsx'):
        return extract_text_from_excel(file_path)
    else:
        raise ValueError("Unsupported file format")
def split_to_dict(text):
    parts = text.split('end')
    question_dict = {}
    for part in parts:
        if part:
            split_part = part.strip().split(' ', 1)
            if len(split_part) == 2:
                key, value = split_part
                question_dict[key] = value
    return question_dict
def calculate_mse(predictions, targets):
    mse = mean_squared_error(predictions, targets)
    return mse
def get_best_score(rouge_scores, similarity_score_1_2, bert_sc, answer_score, bleu_score_pe):
    all_scores = [rouge_scores, similarity_score_1_2, bert_sc, answer_score, bleu_score_pe]
    best_score = max(all_scores)
    return best_score

@login_manager.user_loader
def load_user(user_id):
    return Admin(user_id)

@app.route('/', methods=['GET'])
def home():
    return render_template('home.html')
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # Check the username and password
        username = request.form['username']
        password = request.form['password']
        if username in admins and password == admins[username]['password']:
            user = Admin(username)
            login_user(user)
            session['admin_logged_in'] = True
            return redirect(url_for('image_x'))
        return "Invalid admin credentials"
    return render_template('login.html')
@login_required
def admin_logout():
    logout_user()
    session.pop('admin_logged_in', None)
    return "Admin logged out"
@app.route('/image_x', methods=['GET', 'POST'])
def image_x():
    if request.method == 'POST':
        return "Papers uploaded and processed successfully"
    return render_template('image1.html')
@app.route('/compare_scores', methods=['POST'])
def compare_scores():
    manual_score = float(request.form['manual_score'])
    model_score=float(request.form['model_score'])
    # Calculate the model's score (you need to retrieve it from somewhere)

    # Calculate accuracy
    accuracy = calculate_accuracy(manual_score, model_score)

    return render_template('accuracy.html',model_score=model_score,manual_score=manual_score,accuracy=accuracy)

def calculate_accuracy(manual_score, model_score):
    # Calculate accuracy using your chosen metric (e.g., percentage match)
    accuracy =100-(abs(manual_score - model_score) )
    return accuracy
@app.route('/upload', methods=['POST'])
@login_required
def upload():
    if 'imageInput1' not in request.files or 'imageInput2' not in request.files:
        return "Both images are required"
    
    file1 = request.files['imageInput1']
    file2 = request.files['imageInput2']
    
    if file1.filename == '' or file2.filename == '':
        return "Both files are required"
    
    if file1 and file2:
        filename1 = os.path.join(app.config['UPLOAD_FOLDER'], file1.filename)
        filename2 = os.path.join(app.config['UPLOAD_FOLDER'], file2.filename )
        
        file1.save(filename1)
        file2.save(filename2)
        
    
        if (file1.filename.endswith('.pdf') or file1.filename.endswith('.docx') or file1.filename.endswith('.txt') or file1.filename.endswith('.xlsx') ) and (file2.filename.endswith('.pdf') or file2.filename.endswith('.docx') or file2.filename.endswith('.txt') or file2.filename.endswith('.xlsx')):
            text1 = extract_text_from_file(filename1)
            text2 = extract_text_from_file(filename2)
        elif(file1.filename.endswith('.jpg') or file1.filename.endswith('.jpeg') or file1.filename.endswith('.png')) and (file2.filename.endswith('.jpg') or file2.filename.endswith('.jpeg') or file2.filename.endswith('.png')):
            text1 = pytesseract.image_to_string(filename1)
            text2 = pytesseract.image_to_string(filename2)
        else:
            return "Unsupported file format"
            
        
        text1 = text1.translate(str.maketrans('', '', string.punctuation))
        text2 = text2.translate(str.maketrans('', '', string.punctuation))
        question_key_keywords = set(text2.split())
        answer_score = evaluate_answer_with_keywords(text1, question_key_keywords)

        #spellchecker
        from spellchecker import SpellChecker
        spell = SpellChecker()
        words = text2.split()
        mistake_count=0
        corrected_words = [spell.correction(word) for word in words]
        swords=set(words)
        scorrected_words=set(corrected_words)
        mistakes=(len(list(swords-scorrected_words)))
        reduced_marks=mistakes/5
       #answer sheeet
        text = text1
         
        # tokenization
        words = word_tokenize(text)
        #removing stop words
        stop_words = set(stopwords.words('english'))
        filtered_words = [word for word in words if word.lower() not in stop_words]
        # Join the filtered words back into a sentence
        filtered_text = ' '.join(filtered_words)
        #stemming
        stemmer = PorterStemmer()
        stemmed_words = [stemmer.stem(word) for word in filtered_text]
        stemmed_text = ''.join(stemmed_words)

        #question key sheet
        text = text2
        # tokenization
        words = word_tokenize(text)
        #removing stop words
        stop_words = set(stopwords.words('english'))
        filtered_words = [word for word in words if word.lower() not in stop_words]
        # Join the filtered words back into a sentence
        filtered_text1 = ' '.join(filtered_words)
        #stemming
        stemmer = PorterStemmer()
        stemmed_words = [stemmer.stem(word) for word in filtered_text1]
        stemmed_text1 = ''.join(stemmed_words)

      # You can also use a weighted average if desired
        ngram_similarity = calculate_jaccard_similarity(stemmed_text, stemmed_text1, n=1)
        ngram= ngram_similarity*100

        reference = [stemmed_text]  # Reference text
        candidate = stemmed_text1   # Generated text

        # ROUGE-N scores (1-gram to 4-gram)
        rouge_scores = {}
        for n in range(1, 5):
            rouge_scores[f'rouge-{n}'] = sentence_bleu(reference, candidate, weights=(1.0/n,)*n, smoothing_function=SmoothingFunction().method1)
        rouge_1_score = rouge_scores['rouge-1']
        total=(rouge_1_score*100)
        rouge_scores=total-reduced_marks
        # Calculate n-gram based similarity

        texts = [stemmed_text, stemmed_text1]

# Initialize a CountVectorizer
        vectorizer = CountVectorizer()

# Fit the vectorizer on the texts and transform them into BoW vectors
        bow_matrix = vectorizer.fit_transform(texts)

# Calculate the cosine similarity between all pairs of documents
        cosine_similarities = cosine_similarity(bow_matrix)
        similarity_score_1= (cosine_similarities[0][1])
        similarity_score_1_2 =(similarity_score_1)*100

        model_name = "bert-base-uncased"
        tokenizer = AutoTokenizer.from_pretrained(model_name)
        model = AutoModelForSequenceClassification.from_pretrained(model_name)
        text3 = stemmed_text
        text4 = stemmed_text1

# Tokenize and encode the texts
        inputs = tokenizer(text3, text4, return_tensors="pt", padding=True, truncation=True)
        with torch.no_grad():
          logits = model(**inputs).logits

# Calculate the softmax to get similarity scores
        similarity_scores = torch.softmax(logits, dim=1)

# Extract the similarity score for 'text1' and 'text2'
        score_for_text1 = similarity_scores[0][1].item() 
        bert_sc=score_for_text1 *100
#         # Calculate BLEU score
        reference_tokens = (stemmed_text1)
        candidate_tokens = (stemmed_text)

#     # Calculate BLEU score with smoothing
        smoothing = SmoothingFunction().method4
        bleu_score = sentence_bleu([reference_tokens], candidate_tokens, smoothing_function=smoothing)
#     # Convert BLEU score to percentage (optional)
        bleu_score_pe = bleu_score * 100
        best_score = get_best_score(rouge_scores, similarity_score_1_2, bert_sc, answer_score, bleu_score_pe)
    return render_template('result1.html', text1=text1, text2=text2, answer=stemmed_text, question=stemmed_text1,
                           rouge=rouge_scores, wob=similarity_score_1_2, similarity_b=bert_sc, keywords_score=answer_score,
                           bleu_sc=bleu_score_pe ,best=best_score)


if __name__ == '__main__':
    app.run(debug=True)
